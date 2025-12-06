#!/usr/bin/env python3
"""
Quick script to create a sample Excel file for testing the validation tool.
Requires: pip install openpyxl

Usage: python create-test-excel.py
Output: test-model.xlsx
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl not installed")
    print("Install it with: pip install openpyxl")
    exit(1)

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Create Results sheet
results = wb.create_sheet("Results", 0)
results['A1'] = "Youth Employment Skills Training - Results"
results['A1'].font = Font(bold=True, size=14)

results['A3'] = "Key Metrics"
results['A3'].font = Font(bold=True)

results['A5'] = "SROI (Social Return on Investment)"
results['B5'] = "1:5.5"
results['B5'].font = Font(size=12, color="0000FF")

results['A6'] = "Total Social Value Created"
results['B6'] = "£275,000"

results['A7'] = "Total Investment"
results['B7'] = "£50,000"

results['A9'] = "Participants Helped"
results['B9'] = 150

results['A10'] = "Jobs Secured"
results['B10'] = 98

# Create Assumptions sheet
assumptions = wb.create_sheet("Assumptions", 1)
assumptions['A1'] = "Value Factors & Assumptions"
assumptions['A1'].font = Font(bold=True, size=14)

assumptions['A3'] = "Assumption"
assumptions['B3'] = "Value"
assumptions['C3'] = "Source"
assumptions['A3'].font = Font(bold=True)
assumptions['B3'].font = Font(bold=True)
assumptions['C3'].font = Font(bold=True)

data = [
    ["Average UK Salary", "£33,000", "ONS 2024"],
    ["Wellbeing Value (QALY)", "£8,500", "NHS/HM Treasury"],
    ["Attribution Factor", "60%", "Academic Research"],
    ["Deadweight Factor", "25%", "IES Research"],
    ["Displacement Factor", "10%", "Sector Standards"],
    ["Drop-off per year", "15%", "Longitudinal Studies"],
]

row = 4
for item in data:
    assumptions[f'A{row}'] = item[0]
    assumptions[f'B{row}'] = item[1]
    assumptions[f'C{row}'] = item[2]
    row += 1

# Create Data Inputs sheet
inputs = wb.create_sheet("Data Inputs", 2)
inputs['A1'] = "Data Input Fields"
inputs['A1'].font = Font(bold=True, size=14)

inputs['A3'] = "Field"
inputs['B3'] = "Example Value"
inputs['C3'] = "Mandatory?"
inputs['A3'].font = Font(bold=True)
inputs['B3'].font = Font(bold=True)
inputs['C3'].font = Font(bold=True)

input_data = [
    ["Number of Participants", 150, "YES"],
    ["Program Cost", "£50,000", "YES"],
    ["Duration (months)", 6, "YES"],
    ["Age Range", "16-24", "NO"],
    ["Employment Rate Achieved", "65%", "NO (uses research default)"],
    ["Average Salary of Jobs", "£25,000", "NO (uses UK average)"],
]

row = 4
for item in input_data:
    inputs[f'A{row}'] = item[0]
    inputs[f'B{row}'] = item[1]
    inputs[f'C{row}'] = item[2]

    # Highlight mandatory fields
    if item[2] == "YES":
        inputs[f'C{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    row += 1

# Create Location Data sheet
location = wb.create_sheet("UK Location Data", 3)
location['A1'] = "UK-Specific Data Table"
location['A1'].font = Font(bold=True, size=14)

location['A3'] = "Metric"
location['B3'] = "UK Value"
location['C3'] = "Source/Year"
location['A3'].font = Font(bold=True)
location['B3'].font = Font(bold=True)
location['C3'].font = Font(bold=True)

location_data = [
    ["Youth Unemployment Rate", "11.2%", "ONS 2024"],
    ["Median Salary", "£33,000", "ONS 2024"],
    ["Minimum Wage", "£11.44/hour", "Gov 2024"],
    ["Tax Rate (Basic)", "20%", "HMRC"],
    ["National Insurance", "12%", "HMRC"],
    ["Wellbeing Value", "£8,500", "HM Treasury"],
]

row = 4
for item in location_data:
    location[f'A{row}'] = item[0]
    location[f'B{row}'] = item[1]
    location[f'C{row}'] = item[2]
    row += 1

location['A12'] = "Note: This table can be expanded for other countries"
location['A12'].font = Font(italic=True, color="666666")

# Save workbook
filename = "test-model.xlsx"
wb.save(filename)
print(f"✓ Created {filename}")
print(f"✓ Sheets: {', '.join(wb.sheetnames)}")
print(f"✓ SROI value: 1:5.5 (within acceptable range)")
print("\nYou can now upload this file to the validation tool!")
